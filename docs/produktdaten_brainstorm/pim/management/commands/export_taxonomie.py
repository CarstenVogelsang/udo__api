import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.conf import settings
from django.core.management.base import BaseCommand
from pim.models_taxonomie import Taxonomie, TaxoSortiment, TaxoProduktwelt, TaxoProduktgruppe, TaxoProduktkategorie

class Command(BaseCommand):
    help = "Exportiert die gesamte Taxonomie-Hierarchie für eine angegebene Taxonomie als Excel-Datei."

    def add_arguments(self, parser):
        parser.add_argument("taxonomie_name", type=str, help="Name der Taxonomie, die exportiert werden soll.")

    def handle(self, *args, **kwargs):
        taxonomie_name = kwargs["taxonomie_name"]

        # Taxonomie prüfen
        try:
            taxonomie = Taxonomie.objects.get(name=taxonomie_name)
        except Taxonomie.DoesNotExist:
            self.stdout.write(self.style.ERROR(f"❌ Taxonomie '{taxonomie_name}' nicht gefunden."))
            return

        # Exportverzeichnis vorbereiten
        export_dir = getattr(settings, "EXPORT_DIRECTORY", "export")
        export_dir = os.path.abspath(export_dir)
        os.makedirs(export_dir, exist_ok=True)

        # Excel-Arbeitsmappe erstellen
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Taxonomie_{taxonomie_name}"

        # Kopfzeile
        headers = [
            "Taxonomie.Name",
            "Sortiment.Code", "Sortiment.Name",
            "Produktwelt.Code", "Produktwelt.Name",
            "Produktgruppe.Code", "Produktgruppe.Name",
            "Produktkategorie.Code", "Produktkategorie.Name"
        ]
        ws.append(headers)
        
                # Styling der Kopfzeile
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font

        # Durch alle zugehörigen Sortimente navigieren
        for sortiment in taxonomie.sortimente.all():
            for produktwelt in sortiment.produktwelten.all():
                for produktgruppe in produktwelt.produktgruppen.all():
                    produktkategorien = list(produktgruppe.produktkategorien.all())
                    if produktkategorien:
                        for produktkategorie in produktkategorien:
                            ws.append([
                                taxonomie.name,
                                sortiment.code, sortiment.name_de,
                                produktwelt.code, produktwelt.name_de,
                                produktgruppe.code, produktgruppe.name_de,
                                produktkategorie.code, produktkategorie.name_de
                            ])
                    else:
                        ws.append([
                            taxonomie.name,
                            sortiment.code, sortiment.name_de,
                            produktwelt.code, produktwelt.name_de,
                            produktgruppe.code, produktgruppe.name_de,
                            "", ""
                        ])

        # Spaltenbreiten anpassen (nur für Name-Spalten)
        column_widths = {
            "A": 20,  # Taxonomie.Name
            "B": 20,  # Sortiment.Code
            "C": 30,  # Sortiment.Name
            "D": 20,  # Produktwelt.Code
            "E": 30,  # Produktwelt.Name
            "F": 20,  # Produktgruppe.Code
            "G": 35,  # Produktgruppe.Name
            "H": 20,  # Produktkategorie.Code
            "I": 35   # Produktkategorie.Name
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Datei speichern
        safe_name = re.sub(r'\W+', '_', taxonomie_name)
        filename = os.path.join(export_dir, f"export_taxonomie_{safe_name}.xlsx")
        wb.save(filename)

        self.stdout.write(self.style.SUCCESS(f"✅ Taxonomie exportiert: {filename}"))