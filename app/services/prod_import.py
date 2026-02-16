"""
Product Import Service — imports articles from Excel files.

Hybrid approach:
- Core fields (prod_artikel columns) via ETL field mappings
- EAV properties (prod_artikel_eigenschaft) via code-matching convention
  (Excel column name matches prod_eigenschaft.code)
- Sortiment assignment per import run
"""
import json
import logging
from datetime import datetime
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.etl import (
    EtlFieldMapping,
    EtlImportFile,
    EtlImportLog,
    EtlImportRecord,
    EtlSource,
    EtlTableMapping,
)
from app.models.prod import (
    ProdArtikel,
    ProdArtikelBild,
    ProdArtikelEigenschaft,
    ProdArtikelSortiment,
    ProdArtikelText,
    ProdEigenschaft,
    ProdSortiment,
    ProdWerteliste,
)
from app.services.etl import TRANSFORMS, EtlService

logger = logging.getLogger(__name__)

BATCH_SIZE = 500

# Language column convention: Excel column name → (sprache, target_field)
# These columns are NOT handled via FieldMappings or EAV, but upserted
# into prod_artikel_text per detected language.
SPRACH_SPALTEN: dict[str, tuple[str, str]] = {
    "Bez Englisch": ("en", "bezeichnung"),
    "Bez Französisch": ("fr", "bezeichnung"),
    "Bez Franzoesisch": ("fr", "bezeichnung"),
    "Memo Englisch": ("en", "beschreibung"),
    "Memo Französisch": ("fr", "beschreibung"),
    "Memo Franzoesisch": ("fr", "beschreibung"),
}


class ProdImportService:
    """Import product data from Excel. Core fields via ETL mappings,
    properties via eigenschaft-code matching."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self._etl_service = EtlService(db)
        # Caches
        self._eigenschaft_cache: dict[str, ProdEigenschaft] = {}  # code → model
        self._werteliste_cache: dict[str, set[str]] = {}  # typ → set of valid codes
        self._ean_cache: dict[str, str] = {}  # ean → artikel.id
        self._artnr_cache: dict[str, str] = {}  # "hersteller_id:artnr" → artikel.id
        self._fk_caches: dict[str, dict] = {}  # table.field → {value: uuid}

    async def run_import(
        self,
        source_id: str,
        import_file_id: str,
        sortiment_code: str,
        dry_run: bool = False,
    ) -> dict:
        """Execute a product import from an uploaded Excel file.

        Args:
            source_id: EtlSource UUID
            import_file_id: EtlImportFile UUID
            sortiment_code: Sortiment to assign (e.g. 'moba', 'sammler')
            dry_run: If True, validate only without writing
        """
        batch_id = str(uuid4())
        stats = {
            "batch_id": batch_id,
            "dry_run": dry_run,
            "rows_read": 0,
            "artikel_created": 0,
            "artikel_updated": 0,
            "artikel_skipped": 0,
            "eigenschaften_set": 0,
            "sortiment_assigned": 0,
            "texte_upserted": 0,
            "errors": 0,
            "error_details": [],
        }

        # 1. Load source, table mapping, field mappings
        source = await self._load_source(source_id)
        if not source:
            stats["error_details"].append(f"EtlSource {source_id} nicht gefunden")
            stats["errors"] = 1
            return stats

        table_mapping = await self._get_table_mapping(source_id, "prod_artikel")
        if not table_mapping:
            stats["error_details"].append("Kein TableMapping für prod_artikel gefunden")
            stats["errors"] = 1
            return stats

        field_mappings = await self._get_field_mappings(table_mapping.id)

        # 2. Load sortiment
        sortiment = await self._get_sortiment(sortiment_code)
        if not sortiment:
            stats["error_details"].append(f"Sortiment '{sortiment_code}' nicht gefunden")
            stats["errors"] = 1
            return stats

        # 3. Load import file
        import_file = await self._load_import_file(import_file_id)
        if not import_file or not import_file.file_content:
            stats["error_details"].append(f"Import-Datei {import_file_id} nicht gefunden")
            stats["errors"] = 1
            return stats

        # 4. Build caches
        await self._build_caches()
        blueprint_codes = await self._get_blueprint_codes(sortiment.id)

        # 5. Parse Excel
        rows = self._parse_excel(import_file.file_content)
        stats["rows_read"] = len(rows)

        if not rows:
            stats["error_details"].append("Keine Datenzeilen in der Excel-Datei")
            return stats

        # Determine which Excel columns are mapped to core fields
        mapped_source_fields = set()
        for fm in field_mappings:
            mapped_source_fields.add(fm.source_field.lower())
            if fm.source_field_aliases:
                aliases = fm.source_field_aliases
                if isinstance(aliases, str):
                    try:
                        aliases = json.loads(aliases)
                    except (json.JSONDecodeError, TypeError):
                        aliases = []
                for a in aliases:
                    mapped_source_fields.add(a.strip().lower())

        # 6. Create import log
        import_log = EtlImportLog(
            table_mapping_id=table_mapping.id,
            batch_id=batch_id,
            status="running",
            records_read=len(rows),
        )
        if not dry_run:
            self.db.add(import_log)
            await self.db.flush()

        # 7. Process rows
        for i, row in enumerate(rows):
            try:
                result = await self._process_row(
                    row=row,
                    field_mappings=field_mappings,
                    sortiment=sortiment,
                    blueprint_codes=blueprint_codes,
                    mapped_source_fields=mapped_source_fields,
                    batch_id=batch_id,
                    dry_run=dry_run,
                )
                stats[f"artikel_{result['action']}"] = (
                    stats.get(f"artikel_{result['action']}", 0) + 1
                )
                stats["eigenschaften_set"] += result.get("eigenschaften_set", 0)
                stats["sortiment_assigned"] += result.get("sortiment_assigned", 0)
                stats["texte_upserted"] += result.get("texte_upserted", 0)
            except Exception as e:
                stats["errors"] += 1
                if len(stats["error_details"]) < 20:
                    stats["error_details"].append(f"Zeile {i + 2}: {str(e)}")
                logger.warning(f"Zeile {i + 2}: {e}")

            if not dry_run and (i + 1) % BATCH_SIZE == 0:
                await self.db.commit()
                logger.info(f"Batch commit at row {i + 1}")

        # 8. Finalize
        if not dry_run:
            import_log.finished_at = datetime.utcnow()
            import_log.status = "success" if stats["errors"] == 0 else "partial"
            import_log.records_created = stats["artikel_created"]
            import_log.records_updated = stats["artikel_updated"]
            import_log.records_skipped = stats["artikel_skipped"]
            import_log.records_failed = stats["errors"]
            await self.db.commit()

        return stats

    async def preview_row(
        self,
        source_id: str,
        import_file_id: str,
        row_index: int = 0,
    ) -> dict:
        """Preview a single row transformation (for debugging mappings)."""
        table_mapping = await self._get_table_mapping(source_id, "prod_artikel")
        if not table_mapping:
            return {"error": "Kein TableMapping für prod_artikel"}

        field_mappings = await self._get_field_mappings(table_mapping.id)

        import_file = await self._load_import_file(import_file_id)
        if not import_file or not import_file.file_content:
            return {"error": "Import-Datei nicht gefunden"}

        rows = self._parse_excel(import_file.file_content)
        if row_index >= len(rows):
            return {"error": f"Zeile {row_index} nicht vorhanden (nur {len(rows)} Zeilen)"}

        row = rows[row_index]
        mapped = self._apply_field_mappings(row, field_mappings)

        return {
            "source_row": row,
            "mapped_fields": mapped,
            "unmapped_columns": [
                col for col in row.keys()
                if col.lower() not in {fm.source_field.lower() for fm in field_mappings}
            ],
        }

    # ========== Internal Methods ==========

    async def _process_row(
        self,
        row: dict[str, Any],
        field_mappings: list,
        sortiment: ProdSortiment,
        blueprint_codes: set[str],
        mapped_source_fields: set[str],
        batch_id: str,
        dry_run: bool,
    ) -> dict:
        """Process a single Excel row."""
        result = {"action": "skipped", "eigenschaften_set": 0, "sortiment_assigned": 0, "texte_upserted": 0}

        # 1. Map core fields
        data = self._apply_field_mappings(row, field_mappings)

        # 2. Resolve FK lookups
        await self._resolve_fk_lookups(data, field_mappings, row)

        # 3. Validate required fields
        if not data.get("hersteller_id"):
            raise ValueError("hersteller_id fehlt (kein FK-Mapping oder Lookup gescheitert)")
        if not data.get("marke_id"):
            raise ValueError("marke_id fehlt")
        if not data.get("artikelnummer_hersteller"):
            raise ValueError("artikelnummer_hersteller fehlt")
        if not data.get("bezeichnung"):
            raise ValueError("bezeichnung fehlt")

        # 4. Dedup
        existing_id = self._dedup_artikel(data)

        if dry_run:
            result["action"] = "updated" if existing_id else "created"
            return result

        # 5. Upsert
        if existing_id:
            artikel_id = await self._update_artikel(existing_id, data, field_mappings)
            result["action"] = "updated"
        else:
            artikel_id = await self._create_artikel(data)
            result["action"] = "created"

        # 6. Track in import log
        self.db.add(EtlImportRecord(
            batch_id=batch_id,
            entity_type="prod_artikel",
            entity_id=artikel_id,
            action=result["action"],
        ))

        # 7. Sortiment assignment
        assigned = await self._assign_sortiment(artikel_id, sortiment.id)
        if assigned:
            result["sortiment_assigned"] = 1

        # 8. EAV properties from unmapped columns
        eav_count = await self._match_and_set_eigenschaften(
            artikel_id, row, mapped_source_fields, blueprint_codes
        )
        result["eigenschaften_set"] = eav_count

        # 9. Language columns → prod_artikel_text
        text_count = await self._upsert_sprach_texte(artikel_id, row)
        result["texte_upserted"] = text_count

        return result

    def _parse_excel(self, file_content: bytes) -> list[dict[str, Any]]:
        """Parse Excel file into list of dicts (column_name -> value)."""
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows()
        header_row = next(rows_iter)
        headers = [
            cell.value.strip() if isinstance(cell.value, str) else str(cell.value or "")
            for cell in header_row
        ]

        result = []
        for row in rows_iter:
            row_dict = {}
            for header, cell in zip(headers, row):
                if header:
                    row_dict[header] = cell.value
            if any(v is not None and str(v).strip() for v in row_dict.values()):
                result.append(row_dict)

        wb.close()
        return result

    def _apply_field_mappings(
        self, row: dict[str, Any], field_mappings: list
    ) -> dict[str, Any]:
        """Apply ETL field mappings to extract core article fields."""
        result = {}
        for fm in field_mappings:
            value = self._get_value_with_aliases(
                row, fm.source_field, fm.source_field_aliases
            )

            # Coerce to string
            if value is not None and not isinstance(value, str):
                value = str(value).strip()

            # Default
            if (value is None or not str(value).strip()) and fm.default_value is not None:
                value = fm.default_value

            if value is None or not str(value).strip():
                result[fm.target_field] = None
                continue

            # Apply transform (skip FK transforms, handled separately)
            if fm.transform and not fm.transform.startswith(("fk_lookup", "ref_current")):
                transform_name = fm.transform.split(":")[0]
                if transform_name in TRANSFORMS:
                    value = TRANSFORMS[transform_name](value)

            result[fm.target_field] = value

        return result

    async def _resolve_fk_lookups(
        self, data: dict, field_mappings: list, source_row: dict
    ):
        """Resolve fk_lookup and fk_lookup_or_create transforms."""
        for fm in field_mappings:
            if not fm.transform:
                continue

            if fm.transform.startswith("fk_lookup:"):
                spec = fm.transform.split(":", 1)[1]
                table, field = spec.split(".", 1)
                value = self._get_value_with_aliases(
                    source_row, fm.source_field, fm.source_field_aliases
                )
                # Fall back to default_value (used for _default_ source fields)
                if value is None and fm.default_value:
                    value = fm.default_value
                if value is not None and not isinstance(value, str):
                    value = str(value).strip()
                if not value or not str(value).strip():
                    data[fm.target_field] = None
                    continue

                resolved = await self._etl_service.fk_lookup(
                    str(value).strip(), table, field, self._fk_caches
                )
                data[fm.target_field] = resolved

            elif fm.transform.startswith("fk_lookup_or_create:"):
                spec = fm.transform.split(":", 1)[1]
                table, field = spec.split(".", 1)
                value = self._get_value_with_aliases(
                    source_row, fm.source_field, fm.source_field_aliases
                )
                # Fall back to default_value
                if value is None and fm.default_value:
                    value = fm.default_value
                if value is not None and not isinstance(value, str):
                    value = str(value).strip()
                if not value or not str(value).strip():
                    data[fm.target_field] = None
                    continue

                resolved = await self._etl_service.fk_lookup_or_create(
                    str(value).strip(), table, field, self._fk_caches
                )
                data[fm.target_field] = resolved

    @staticmethod
    def _get_value_with_aliases(
        row: dict[str, Any], source_field: str, aliases: str | list | None
    ) -> Any:
        """Get value from row, trying source_field first, then aliases."""
        value = row.get(source_field)
        if value is not None:
            return value
        if not aliases:
            return None
        if isinstance(aliases, str):
            try:
                aliases = json.loads(aliases)
            except (json.JSONDecodeError, TypeError):
                return None
        for alias in aliases:
            value = row.get(alias.strip())
            if value is not None:
                return value
        return None

    def _dedup_artikel(self, data: dict) -> str | None:
        """Find existing article by (hersteller_id, artikelnummer_hersteller) or EAN."""
        hersteller_id = data.get("hersteller_id")
        artnr = data.get("artikelnummer_hersteller")

        if hersteller_id and artnr:
            cache_key = f"{hersteller_id}:{artnr}"
            if cache_key in self._artnr_cache:
                return self._artnr_cache[cache_key]

        ean = data.get("ean_gtin")
        if ean and ean.strip() in self._ean_cache:
            return self._ean_cache[ean.strip()]

        return None

    async def _create_artikel(self, data: dict) -> str:
        """Create a new article and update caches."""
        # Remove None FK fields that could cause constraint violations
        clean = {k: v for k, v in data.items() if v is not None or k not in (
            "serie_id", "kategorie_id"
        )}

        artikel = ProdArtikel(**clean)
        self.db.add(artikel)
        await self.db.flush()

        artikel_id = str(artikel.id)

        # Update caches
        if data.get("hersteller_id") and data.get("artikelnummer_hersteller"):
            key = f"{data['hersteller_id']}:{data['artikelnummer_hersteller']}"
            self._artnr_cache[key] = artikel_id
        if data.get("ean_gtin"):
            self._ean_cache[data["ean_gtin"].strip()] = artikel_id

        return artikel_id

    async def _update_artikel(
        self, existing_id: str, data: dict, field_mappings: list
    ) -> str:
        """Update an existing article respecting update rules."""
        result = await self.db.execute(
            select(ProdArtikel).where(ProdArtikel.id == existing_id)
        )
        artikel = result.scalar_one_or_none()
        if not artikel:
            return await self._create_artikel(data)

        rules = {fm.target_field: (fm.update_rule or "always") for fm in field_mappings}
        updated = False

        for field, value in data.items():
            if field in ("id", "erstellt_am", "aktualisiert_am", "geloescht_am"):
                continue
            rule = rules.get(field, "always")
            current = getattr(artikel, field, None)

            if rule == "never":
                continue
            elif rule == "if_empty":
                if current is not None and str(current).strip():
                    continue

            if value != current:
                setattr(artikel, field, value)
                updated = True

        if updated:
            artikel.aktualisiert_am = datetime.utcnow()
            await self.db.flush()

        return str(artikel.id)

    async def _assign_sortiment(self, artikel_id: str, sortiment_id: str) -> bool:
        """Assign sortiment if not already assigned."""
        existing = await self.db.execute(
            select(ProdArtikelSortiment).where(
                ProdArtikelSortiment.artikel_id == artikel_id,
                ProdArtikelSortiment.sortiment_id == sortiment_id,
            )
        )
        if existing.scalar_one_or_none():
            return False

        self.db.add(ProdArtikelSortiment(
            artikel_id=artikel_id,
            sortiment_id=sortiment_id,
        ))
        await self.db.flush()
        return True

    async def _match_and_set_eigenschaften(
        self,
        artikel_id: str,
        row: dict[str, Any],
        mapped_source_fields: set[str],
        blueprint_codes: set[str],
    ) -> int:
        """Match unmapped Excel columns to eigenschaft definitions by code.

        Convention: Excel column name (case-insensitive) == eigenschaft.code
        """
        count = 0

        for col_name, value in row.items():
            if value is None:
                continue
            col_lower = col_name.strip().lower()

            # Skip columns that are mapped to core fields
            if col_lower in mapped_source_fields:
                continue

            # Check if this column matches an eigenschaft code
            eigenschaft = self._eigenschaft_cache.get(col_lower)
            if not eigenschaft:
                continue

            # Optionally check if eigenschaft is in the sortiment blueprint
            if blueprint_codes and eigenschaft.code not in blueprint_codes:
                continue

            # Determine value column based on daten_typ
            wert_data = self._prepare_eigenschaft_value(eigenschaft, value)
            if wert_data is None:
                continue

            # Upsert eigenschaft value
            existing = await self.db.execute(
                select(ProdArtikelEigenschaft).where(
                    ProdArtikelEigenschaft.artikel_id == artikel_id,
                    ProdArtikelEigenschaft.eigenschaft_id == eigenschaft.id,
                )
            )
            ae = existing.scalar_one_or_none()

            if ae:
                ae.wert_text = wert_data.get("wert_text")
                ae.wert_ganzzahl = wert_data.get("wert_ganzzahl")
                ae.wert_dezimal = wert_data.get("wert_dezimal")
                ae.wert_bool = wert_data.get("wert_bool")
            else:
                self.db.add(ProdArtikelEigenschaft(
                    artikel_id=artikel_id,
                    eigenschaft_id=eigenschaft.id,
                    **wert_data,
                ))

            count += 1

        if count > 0:
            await self.db.flush()

        return count

    async def _upsert_sprach_texte(
        self,
        artikel_id: str,
        row: dict[str, Any],
    ) -> int:
        """Detect language columns and upsert into prod_artikel_text.

        Groups columns by language, then upserts one ProdArtikelText
        record per language (bezeichnung + beschreibung).
        """
        # Collect values per language: {sprache: {field: value}}
        lang_data: dict[str, dict[str, str]] = {}

        for col_name, value in row.items():
            if value is None or not str(value).strip():
                continue
            mapping = SPRACH_SPALTEN.get(col_name)
            if not mapping:
                continue
            sprache, field = mapping
            text_value = str(value).strip().replace("_x000D_", "")
            if text_value and text_value != "nan":
                lang_data.setdefault(sprache, {})[field] = text_value

        if not lang_data:
            return 0

        count = 0
        for sprache, fields in lang_data.items():
            existing = await self.db.execute(
                select(ProdArtikelText).where(
                    ProdArtikelText.artikel_id == artikel_id,
                    ProdArtikelText.sprache == sprache,
                )
            )
            text_record = existing.scalar_one_or_none()

            if text_record:
                for field, value in fields.items():
                    setattr(text_record, field, value)
            else:
                self.db.add(ProdArtikelText(
                    artikel_id=artikel_id,
                    sprache=sprache,
                    **fields,
                ))
            count += 1

        if count > 0:
            await self.db.flush()

        return count

    def _prepare_eigenschaft_value(
        self, eigenschaft: ProdEigenschaft, raw_value: Any
    ) -> dict | None:
        """Convert raw Excel value to typed EAV value dict."""
        dt = eigenschaft.daten_typ
        value_str = str(raw_value).strip() if raw_value is not None else ""

        if not value_str:
            return None

        if dt == "werteliste":
            # Validate against werteliste
            valid_codes = self._werteliste_cache.get(eigenschaft.werteliste_typ, set())
            code_upper = value_str.upper()
            if code_upper not in valid_codes and value_str not in valid_codes:
                logger.warning(
                    f"Werteliste '{eigenschaft.werteliste_typ}': "
                    f"'{value_str}' ungültig, überspringe"
                )
                return None
            return {"wert_text": value_str}

        elif dt == "text":
            return {"wert_text": value_str}

        elif dt == "ganzzahl":
            try:
                return {"wert_ganzzahl": int(float(value_str))}
            except (ValueError, TypeError):
                return None

        elif dt == "dezimal":
            try:
                return {"wert_dezimal": float(value_str)}
            except (ValueError, TypeError):
                return None

        elif dt == "bool":
            truthy = value_str.lower() in ("true", "1", "ja", "yes", "x")
            return {"wert_bool": truthy}

        return {"wert_text": value_str}

    # ========== Cache Building ==========

    async def _build_caches(self):
        """Pre-load caches for dedup and eigenschaft matching."""
        # Eigenschaft cache: code → model
        result = await self.db.execute(select(ProdEigenschaft))
        for e in result.scalars().all():
            self._eigenschaft_cache[e.code.lower()] = e

        # Werteliste cache: typ → set of valid codes
        result = await self.db.execute(
            select(ProdWerteliste).where(ProdWerteliste.ist_aktiv == True)  # noqa: E712
        )
        for w in result.scalars().all():
            self._werteliste_cache.setdefault(w.typ, set()).add(w.code)

        # Artikel dedup caches
        result = await self.db.execute(
            select(
                ProdArtikel.id,
                ProdArtikel.hersteller_id,
                ProdArtikel.artikelnummer_hersteller,
                ProdArtikel.ean_gtin,
            ).where(ProdArtikel.geloescht_am == None)  # noqa: E711
        )
        for row in result.all():
            artikel_id = str(row.id)
            if row.hersteller_id and row.artikelnummer_hersteller:
                key = f"{row.hersteller_id}:{row.artikelnummer_hersteller}"
                self._artnr_cache[key] = artikel_id
            if row.ean_gtin:
                self._ean_cache[row.ean_gtin.strip()] = artikel_id

    async def _get_blueprint_codes(self, sortiment_id: str) -> set[str]:
        """Get eigenschaft codes for a sortiment blueprint."""
        from app.models.prod import ProdSortimentEigenschaft

        result = await self.db.execute(
            select(ProdSortimentEigenschaft)
            .where(ProdSortimentEigenschaft.sortiment_id == sortiment_id)
        )
        codes = set()
        for se in result.scalars().all():
            eigenschaft = self._eigenschaft_cache.get(
                next((k for k, v in self._eigenschaft_cache.items()
                      if v.id == se.eigenschaft_id), None)
            )
            if eigenschaft:
                codes.add(eigenschaft.code)
        return codes

    # ========== Loaders ==========

    async def _load_source(self, source_id: str) -> EtlSource | None:
        result = await self.db.execute(
            select(EtlSource).where(EtlSource.id == source_id)
        )
        return result.scalar_one_or_none()

    async def _get_table_mapping(
        self, source_id: str, target_table: str
    ) -> EtlTableMapping | None:
        result = await self.db.execute(
            select(EtlTableMapping).where(
                EtlTableMapping.source_id == source_id,
                EtlTableMapping.target_table == target_table,
            )
        )
        return result.scalar_one_or_none()

    async def _get_field_mappings(self, table_mapping_id: str) -> list:
        result = await self.db.execute(
            select(EtlFieldMapping)
            .where(EtlFieldMapping.table_mapping_id == table_mapping_id)
            .order_by(EtlFieldMapping.erstellt_am)
        )
        return list(result.scalars().all())

    async def _load_import_file(self, import_file_id: str) -> EtlImportFile | None:
        result = await self.db.execute(
            select(EtlImportFile).where(EtlImportFile.id == import_file_id)
        )
        return result.scalar_one_or_none()

    async def _get_sortiment(self, code: str) -> ProdSortiment | None:
        result = await self.db.execute(
            select(ProdSortiment).where(ProdSortiment.code == code)
        )
        return result.scalar_one_or_none()
