"""
Excel Import Service for Unternehmen + Kontakt data.

Processes Excel files using ETL configuration (EtlSource with connection_type="excel").
Features:
- Multi-stage deduplication with cascading priority
- Field-level update rules (always, if_empty, never)
- Multiple external IDs per entity
- GeoOrt resolution by PLZ
- Batch commits for large imports
"""
import json
import logging
import uuid
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.com import (
    ComUnternehmen, ComKontakt, ComExternalId, ComUnternehmenIdentifikation,
)
from app.models.etl import (
    EtlSource, EtlTableMapping, EtlFieldMapping, EtlImportLog, EtlImportRecord,
)
from app.models.geo import Base, generate_uuid
from app.services.etl import TRANSFORMS, ROW_TRANSFORMS, EtlService

logger = logging.getLogger(__name__)

BATCH_SIZE = 500

# Generic email domains that should NOT be used for deduplication.
# Two different companies may both use gmail.com — that doesn't make them the same.
GENERIC_EMAIL_DOMAINS = {
    "gmail.com", "yahoo.com", "yahoo.de", "hotmail.com", "hotmail.de",
    "outlook.com", "outlook.de", "live.com", "live.de",
    "web.de", "gmx.de", "gmx.net", "gmx.at", "gmx.ch",
    "t-online.de", "aol.com", "icloud.com", "me.com",
    "freenet.de", "arcor.de", "online.de", "telekom.de",
    "mail.de", "email.de", "posteo.de", "protonmail.com",
    "yandex.com", "zoho.com",
}

# Transforms that produce valid output from None/empty values.
# These run BEFORE the default_value/None short-circuit so they always execute.
NULL_SAFE_TRANSFORMS = {
    "x_to_bool", "invert_x_flag", "map_store_typ",
    "map_bonitaet_score", "map_loeschkennzeichen",
}

# Implicit FK resolution: target fields that need UUID lookup but don't use
# explicit fk_lookup: transforms. Maps target_field -> (table, lookup_field, filter).
# Optional 3rd element is a SQL WHERE clause for filtered lookups.
IMPLICIT_FK_MAP = {
    "marke_id": ("com_marke", "name"),
    "serie_id": ("com_serie", "name"),
    "dienstleistung_id": ("com_dienstleistung", "name"),
    "sprache_id": ("bas_sprache", "code"),
    "status_id": ("bas_status", "code", "kontext = 'unternehmen'"),
}


class ExcelImportService:
    """Processes Excel file imports for Unternehmen + Kontakt data."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self._etl_service = EtlService(db)
        self._batch_id: str = ""
        # FK lookup caches (shared with EtlService for fk_lookup_or_create)
        self._fk_caches: dict[str, dict] = {}
        # Dedup caches (loaded once, O(1) lookup per row)
        self._plz_cache: dict[str, str] = {}         # normalized PLZ -> geo_ort.id
        self._email_cache: dict[str, str] = {}        # normalized email -> unternehmen.id
        self._url_cache: dict[str, str] = {}          # normalized url -> unternehmen.id
        self._phone_cache: dict[str, str] = {}        # normalized phone -> unternehmen.id
        self._address_cache: dict[str, str] = {}      # "plz|strasse" -> unternehmen.id
        self._name_cache: dict[str, str] = {}         # normalized firmierung -> unternehmen.id
        self._extid_cache: dict[str, str] = {}        # "source:type:value" -> entity_id
        self._entity_extid_set: set[str] = set()     # "entity_type:entity_id:source:type"
        self._ident_cache: dict[str, str] = {}        # "typ:wert" -> unternehmen_id

    async def run_import(
        self,
        source_name: str,
        file_content: bytes,
        dry_run: bool = False,
    ) -> dict:
        """
        Main entry point: Parse Excel, apply mappings, dedup, upsert.

        Returns dict with import statistics.
        """
        batch_id = str(uuid.uuid4())
        self._batch_id = batch_id

        # 1. Load ETL configuration
        source, u_mapping, k_mapping, u_fields, k_fields, extid_fields, bizid_fields, junction_configs = (
            await self._load_config(source_name)
        )

        # 2. Parse Excel
        rows = self._parse_excel(file_content)
        logger.info(f"Excel parsed: {len(rows)} rows")

        # Collect all field mappings for FK cache building
        all_fields = u_fields + k_fields
        for _jm, jf in junction_configs:
            all_fields += jf

        # 3. Build dedup caches
        await self._build_dedup_caches()
        await self._build_plz_cache()
        await self._build_fk_caches(all_fields)

        # 4. Create import logs
        u_log = EtlImportLog(
            table_mapping_id=u_mapping.id,
            batch_id=batch_id,
            status="running",
            records_read=len(rows),
        )
        k_log = EtlImportLog(
            table_mapping_id=k_mapping.id,
            batch_id=batch_id,
            status="running",
            records_read=len(rows),
        )
        if not dry_run:
            self.db.add(u_log)
            self.db.add(k_log)
            await self.db.flush()

        # 5. Process rows
        stats = {
            "rows_read": len(rows),
            "unternehmen_created": 0,
            "unternehmen_updated": 0,
            "unternehmen_skipped": 0,
            "kontakte_created": 0,
            "kontakte_updated": 0,
            "kontakte_skipped": 0,
            "junction_created": 0,
            "junction_skipped": 0,
            "junction_details": {},
            "errors": 0,
            "error_details": [],
        }

        for i, row in enumerate(rows):
            try:
                await self._process_row(
                    row, i + 2,  # Excel row number (1-based header + 1-based data)
                    u_fields, k_fields, extid_fields, bizid_fields,
                    junction_configs,
                    source_name, stats, dry_run,
                )
            except Exception as e:
                stats["errors"] += 1
                msg = f"Zeile {i + 2}: {str(e)}"
                if len(stats["error_details"]) < 20:
                    stats["error_details"].append(msg)
                logger.warning(msg)

            # Batch commit
            if not dry_run and (i + 1) % BATCH_SIZE == 0:
                await self.db.commit()
                logger.info(f"Batch commit at row {i + 1}")

        # 6. Final commit + update logs
        if not dry_run:
            u_log.finished_at = datetime.utcnow()
            u_log.status = "success"
            u_log.records_created = stats["unternehmen_created"]
            u_log.records_updated = stats["unternehmen_updated"]
            u_log.records_skipped = stats["unternehmen_skipped"]
            u_log.records_failed = stats["errors"]

            k_log.finished_at = datetime.utcnow()
            k_log.status = "success"
            k_log.records_created = stats["kontakte_created"]
            k_log.records_updated = stats["kontakte_updated"]
            k_log.records_skipped = stats["kontakte_skipped"]

            await self.db.commit()

        stats["batch_id"] = batch_id
        stats["dry_run"] = dry_run
        return stats

    # ============ Preview (single row, no DB writes) ============

    async def preview_row(self, source_name: str, source_row: dict[str, Any]) -> dict:
        """Preview transform results for a single row without database writes.

        Applies all field mappings, transforms, and FK lookups for every target
        table, then returns a structured result showing source → transform → target
        for each field.
        """
        # 1. Load config
        (source, u_mapping, k_mapping, u_fields, k_fields,
         extid_fields, bizid_fields, junction_configs) = await self._load_config(source_name)

        # 2. Collect all field mappings and build FK caches
        all_fields = u_fields + k_fields
        for _jm, jf in junction_configs:
            all_fields += jf
        await self._build_fk_caches(all_fields)

        # 3. Build table groups: [(table_name, field_mappings, extid_fields, bizid_fields)]
        table_groups: list[tuple[str, list[EtlFieldMapping], list, list]] = [
            ("com_unternehmen", u_fields, extid_fields, bizid_fields),
            ("com_kontakt", k_fields, [], []),
        ]
        for jm, jf in junction_configs:
            table_groups.append((jm.target_table, jf, [], []))

        # 4. Process each table
        tables_result = {}
        for table_name, fields, ext_fields, biz_fields in table_groups:
            field_results = []

            # Regular field mappings
            transformed = self._apply_field_mappings(source_row, fields)
            # FK resolution (modifies transformed in-place)
            await self._resolve_fk_lookups(transformed, fields)
            self._resolve_implicit_fks(transformed, fields)

            for fm in fields:
                source_value = self._get_value_with_aliases(
                    source_row, fm.source_field, fm.source_field_aliases
                )
                target_value = transformed.get(fm.target_field)

                # Build label for FK values (reverse-lookup)
                target_label = None
                if target_value and fm.target_field.endswith("_id"):
                    target_label = self._reverse_fk_label(fm.target_field, target_value)

                # Determine status
                source_empty = (
                    source_value is None
                    or (isinstance(source_value, str) and not source_value.strip())
                )
                has_output = target_value is not None and str(target_value).strip() != ""

                if fm.transform and fm.transform.startswith("ref_current:"):
                    status = "placeholder"
                    target_value = f"[ref_current:{fm.transform.split(':', 1)[1]}]"
                elif source_empty and has_output:
                    status = "default"
                elif source_empty and not has_output:
                    status = "empty"
                elif not source_empty and has_output:
                    status = "transformed"
                else:
                    status = "skipped"

                field_results.append({
                    "source_field": fm.source_field,
                    "source_value": source_value,
                    "transform": fm.transform or None,
                    "target_field": fm.target_field,
                    "target_value": target_value,
                    "target_label": target_label,
                    "update_rule": fm.update_rule or "always",
                    "status": status,
                })

            # External ID fields
            for src_field, ext_source, ext_type, *rest in ext_fields:
                aliases = rest[0] if rest else None
                val = self._get_value_with_aliases(source_row, src_field, aliases)
                field_results.append({
                    "source_field": src_field,
                    "source_value": val,
                    "transform": f"external_id:{ext_source}.{ext_type}",
                    "target_field": "_extid",
                    "target_value": f"[ExtID: {ext_source}.{ext_type} = {val}]" if val else None,
                    "target_label": None,
                    "update_rule": "always",
                    "status": "placeholder" if val else "empty",
                })

            # Business ID fields
            for src_field, ident_type, *rest in biz_fields:
                aliases = rest[0] if rest else None
                val = self._get_value_with_aliases(source_row, src_field, aliases)
                field_results.append({
                    "source_field": src_field,
                    "source_value": val,
                    "transform": f"business_id:{ident_type}",
                    "target_field": "_bizid",
                    "target_value": f"[BizID: {ident_type} = {val}]" if val else None,
                    "target_label": None,
                    "update_rule": "always",
                    "status": "placeholder" if val else "empty",
                })

            tables_result[table_name] = {"fields": field_results}

        return {"tables": tables_result}

    def _reverse_fk_label(self, target_field: str, uuid_value: str) -> str | None:
        """Reverse-lookup a UUID to find a human-readable label."""
        # Search all FK caches for a matching UUID
        for cache_key, cache in self._fk_caches.items():
            for label, cached_id in cache.items():
                if str(cached_id) == str(uuid_value):
                    return str(label)
        return None

    async def _process_row(
        self,
        row: dict[str, Any],
        row_num: int,
        u_fields: list[EtlFieldMapping],
        k_fields: list[EtlFieldMapping],
        extid_fields: list[tuple[str, str, str]],  # [(source_field, source_name, id_type)]
        bizid_fields: list[tuple[str, str]],        # [(source_field, ident_type)]
        junction_configs: list[tuple[EtlTableMapping, list[EtlFieldMapping]]],
        source_name: str,
        stats: dict,
        dry_run: bool,
    ):
        """Process a single Excel row: Unternehmen + Kontakt + Junction Tables."""
        # Apply Unternehmen field mappings
        u_data = self._apply_field_mappings(row, u_fields)

        # Resolve fk_lookup_or_create transforms (e.g., Organisation)
        await self._resolve_fk_or_create(u_data, u_fields, row)

        # Resolve GeoOrt by PLZ
        if "geo_ort_id" in u_data and u_data["geo_ort_id"]:
            plz = u_data["geo_ort_id"]  # At this point it's the normalized PLZ string
            u_data["geo_ort_id"] = self._plz_cache.get(plz)

        # Resolve implicit FKs (e.g., sprache_id code -> UUID)
        self._resolve_implicit_fks(u_data, u_fields)

        # Skip if no meaningful data
        if not u_data.get("kurzname") and not u_data.get("firmierung"):
            stats["unternehmen_skipped"] += 1
            stats["kontakte_skipped"] += 1
            return

        # Dedup Unternehmen
        existing_id, match_method = self._dedup_unternehmen(
            u_data, row, source_name, extid_fields, bizid_fields,
        )

        if dry_run:
            if existing_id:
                stats["unternehmen_updated"] += 1
                unternehmen_id = existing_id
            else:
                stats["unternehmen_created"] += 1
                unternehmen_id = str(generate_uuid())  # Dummy ID for ref_context

            # Process junctions in dry_run mode
            if junction_configs:
                ref_context = {"id": unternehmen_id, **u_data}
                for jt_mapping, jt_fields in junction_configs:
                    await self._process_junction_row(
                        row, jt_mapping, jt_fields, ref_context, stats,
                        dry_run=True,
                    )

            # Check if kontakt would be created
            k_data = self._apply_field_mappings(row, k_fields)
            if k_data.get("vorname") and k_data.get("nachname"):
                stats["kontakte_created"] += 1
            else:
                stats["kontakte_skipped"] += 1
            return

        # Upsert Unternehmen
        unternehmen_id, u_action = await self._upsert_unternehmen(
            existing_id, u_data, u_fields,
        )
        stats[f"unternehmen_{u_action}"] += 1

        # Track import record
        self.db.add(EtlImportRecord(
            batch_id=self._batch_id,
            entity_type="unternehmen",
            entity_id=unternehmen_id,
            action=u_action,
        ))

        # Update dedup caches with new/updated data
        self._update_dedup_caches(unternehmen_id, u_data)

        # Save external IDs
        await self._save_external_ids(
            "unternehmen", unternehmen_id, row, extid_fields,
        )

        # Save business identifiers (USt-ID, DUNS, etc.)
        await self._save_business_ids(unternehmen_id, row, bizid_fields)

        # Process junction table mappings (e.g., com_unternehmen_organisation)
        if junction_configs:
            ref_context = {"id": unternehmen_id, **u_data}
            for jt_mapping, jt_fields in junction_configs:
                await self._process_junction_row(
                    row, jt_mapping, jt_fields, ref_context, stats,
                )

        # Apply Kontakt field mappings
        k_data = self._apply_field_mappings(row, k_fields)
        await self._resolve_fk_or_create(k_data, k_fields, row)

        # Skip kontakt if no name
        if not k_data.get("vorname") or not k_data.get("nachname"):
            stats["kontakte_skipped"] += 1
            return

        # Upsert Kontakt
        kontakt_id, k_action = await self._upsert_kontakt(
            unternehmen_id, k_data, k_fields,
        )
        stats[f"kontakte_{k_action}"] += 1

        # Track kontakt import record
        self.db.add(EtlImportRecord(
            batch_id=self._batch_id,
            entity_type="kontakt",
            entity_id=kontakt_id,
            action=k_action,
        ))

    # ============ Alias Helper ============

    @staticmethod
    def _get_value_with_aliases(
        row: dict[str, Any], source_field: str, aliases_json: str | list | None
    ) -> Any:
        """Get value from row, trying source_field first, then aliases."""
        value = row.get(source_field)
        if value is not None:
            return value
        if not aliases_json:
            return None
        if isinstance(aliases_json, str):
            try:
                aliases = json.loads(aliases_json)
            except (json.JSONDecodeError, TypeError):
                return None
        else:
            aliases = aliases_json or []
        for alias in aliases:
            value = row.get(alias.strip())
            if value is not None:
                return value
        return None

    # ============ Excel Parsing ============

    def _parse_excel(self, file_content: bytes) -> list[dict[str, Any]]:
        """Parse Excel file into list of dicts (column_name -> value)."""
        wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows()
        header_row = next(rows_iter)
        headers = [
            cell.value.strip() if isinstance(cell.value, str) else str(cell.value or "")
            for cell in header_row
        ]

        result = []
        for row in rows_iter:
            row_dict = {}
            for header, cell in zip(headers, row):
                if header:
                    row_dict[header] = cell.value
            # Skip completely empty rows
            if any(v is not None and str(v).strip() for v in row_dict.values()):
                result.append(row_dict)

        wb.close()
        return result

    # ============ Config Loading ============

    async def _load_config(self, source_name: str):
        """Load ETL configuration for the given source.

        Returns: (source, u_mapping, k_mapping, u_fields, k_fields, extid_fields,
                  junction_configs)
        where junction_configs is a list of (EtlTableMapping, list[EtlFieldMapping])
        for any additional target tables (e.g., com_unternehmen_organisation).
        """
        result = await self.db.execute(
            select(EtlSource)
            .where(EtlSource.name == source_name)
            .where(EtlSource.is_active.is_(True))
            .where(EtlSource.connection_type == "excel")
        )
        source = result.scalar_one_or_none()
        if not source:
            raise ValueError(f"Excel-Import-Quelle '{source_name}' nicht gefunden oder inaktiv.")

        # Load table mappings with field mappings
        result = await self.db.execute(
            select(EtlTableMapping)
            .where(EtlTableMapping.source_id == source.id)
            .where(EtlTableMapping.is_active.is_(True))
        )
        mappings = result.scalars().all()

        u_mapping = None
        k_mapping = None
        other_mappings = []
        for m in mappings:
            if m.target_table == "com_unternehmen":
                u_mapping = m
            elif m.target_table == "com_kontakt":
                k_mapping = m
            else:
                other_mappings.append(m)

        if not u_mapping:
            raise ValueError(f"Kein TableMapping für com_unternehmen in '{source_name}'.")
        if not k_mapping:
            raise ValueError(f"Kein TableMapping für com_kontakt in '{source_name}'.")

        # Load field mappings
        u_fields_result = await self.db.execute(
            select(EtlFieldMapping)
            .where(EtlFieldMapping.table_mapping_id == u_mapping.id)
        )
        k_fields_result = await self.db.execute(
            select(EtlFieldMapping)
            .where(EtlFieldMapping.table_mapping_id == k_mapping.id)
        )

        u_fields_all = list(u_fields_result.scalars().all())
        k_fields_all = list(k_fields_result.scalars().all())

        # Separate external_id and business_id fields from regular fields
        u_fields = []
        extid_fields = []
        bizid_fields = []  # [(source_field, ident_type, aliases)]
        for f in u_fields_all:
            if f.transform and f.transform.startswith("external_id:"):
                # Parse "external_id:smartmail.subscriber_id"
                parts = f.transform.split(":", 1)[1]
                src_name, id_type = parts.split(".", 1)
                extid_fields.append((f.source_field, src_name, id_type, f.source_field_aliases))
            elif f.transform and f.transform.startswith("business_id:"):
                # Parse "business_id:ust_id"
                ident_type = f.transform.split(":", 1)[1]
                bizid_fields.append((f.source_field, ident_type, f.source_field_aliases))
            else:
                u_fields.append(f)

        # Load field mappings for junction/additional tables
        junction_configs = []
        for jm in other_mappings:
            jf_result = await self.db.execute(
                select(EtlFieldMapping)
                .where(EtlFieldMapping.table_mapping_id == jm.id)
            )
            jf_fields = list(jf_result.scalars().all())
            if jf_fields:
                junction_configs.append((jm, jf_fields))

        return (source, u_mapping, k_mapping, u_fields, k_fields_all,
                extid_fields, bizid_fields, junction_configs)

    # ============ Field Mapping ============

    def _apply_field_mappings(
        self,
        row: dict[str, Any],
        field_mappings: list[EtlFieldMapping],
        ref_context: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Apply all field mappings + transformations to a single row."""
        result = {}
        for fm in field_mappings:
            # ref_current: value comes from context, not from Excel row
            if fm.transform and fm.transform.startswith("ref_current:"):
                field = fm.transform.split(":", 1)[1]
                result[fm.target_field] = ref_context.get(field) if ref_context else None
                continue

            value = self._get_value_with_aliases(
                row, fm.source_field, fm.source_field_aliases
            )

            # Coerce to string (openpyxl returns int/float for numeric cells)
            if value is not None and not isinstance(value, str):
                value = str(value).strip()

            # NULL-safe transforms: run even when value is None/empty
            if fm.transform:
                transform_name = fm.transform.split(":")[0]
                if transform_name in NULL_SAFE_TRANSFORMS:
                    if transform_name in TRANSFORMS:
                        result[fm.target_field] = TRANSFORMS[transform_name](value)
                    elif transform_name in ROW_TRANSFORMS:
                        transform_params = (
                            fm.transform.split(":", 1)[1] if ":" in fm.transform else None
                        )
                        result[fm.target_field] = ROW_TRANSFORMS[transform_name](
                            value, row=row, params=transform_params
                        )
                    continue

            # Apply default if value is None/empty
            if value is None or not value.strip():
                if fm.default_value is not None:
                    value = fm.default_value
                elif fm.is_required:
                    continue  # Skip required fields with no value
                else:
                    result[fm.target_field] = None
                    continue

            # Apply transformation
            if fm.transform and fm.transform in TRANSFORMS:
                value = TRANSFORMS[fm.transform](value)
            elif fm.transform and fm.transform.startswith("fk_lookup:"):
                pass  # Handled separately (geo_ort_id by PLZ)
            elif fm.transform and fm.transform.startswith("fk_lookup_or_create:"):
                pass  # Resolved async in _resolve_fk_or_create()
            elif fm.transform:
                # Check for ROW_TRANSFORMS (syntax: "transform_name:param")
                transform_name = fm.transform.split(":")[0]
                transform_params = fm.transform.split(":", 1)[1] if ":" in fm.transform else None
                if transform_name in ROW_TRANSFORMS:
                    value = ROW_TRANSFORMS[transform_name](
                        value, row=row, params=transform_params
                    )

            result[fm.target_field] = value

        return result

    async def _resolve_fk_or_create(
        self,
        data: dict[str, Any],
        field_mappings: list[EtlFieldMapping],
        source_row: dict[str, Any],
    ):
        """Resolve fk_lookup_or_create transforms — creates missing FK records."""
        for fm in field_mappings:
            if not fm.transform or not fm.transform.startswith("fk_lookup_or_create:"):
                continue

            # Parse "fk_lookup_or_create:com_organisation.kurzname"
            lookup_spec = fm.transform.split(":", 1)[1]
            if "." not in lookup_spec:
                continue

            table, field = lookup_spec.split(".", 1)
            value = self._get_value_with_aliases(
                source_row, fm.source_field, fm.source_field_aliases
            )

            # Coerce to string
            if value is not None and not isinstance(value, str):
                value = str(value).strip()
            if not value or not value.strip():
                data[fm.target_field] = None
                continue

            # Lookup or create via EtlService
            resolved_id = await self._etl_service.fk_lookup_or_create(
                value.strip(), table, field, self._fk_caches,
            )
            data[fm.target_field] = resolved_id

    async def _resolve_fk_lookups(
        self,
        data: dict[str, Any],
        field_mappings: list[EtlFieldMapping],
    ):
        """Resolve fk_lookup: transforms (lookup only, no auto-create)."""
        for fm in field_mappings:
            if not fm.transform or not fm.transform.startswith("fk_lookup:"):
                continue
            spec = fm.transform.split(":", 1)[1]
            if "." not in spec:
                continue
            table, field = spec.split(".", 1)
            cache_key = f"{table}.{field}"
            value = data.get(fm.target_field)
            if value is None or (isinstance(value, str) and not value.strip()):
                data[fm.target_field] = None
                continue
            # Cache lookup
            cache = self._fk_caches.get(cache_key, {})
            lookup_val = str(value).strip()
            if lookup_val in cache:
                data[fm.target_field] = str(cache[lookup_val])
            else:
                # DB fallback
                query = text(f"SELECT id FROM {table} WHERE {field} = :val LIMIT 1")
                result = await self.db.execute(query, {"val": lookup_val})
                row = result.fetchone()
                if row:
                    resolved_id = str(row[0])
                    data[fm.target_field] = resolved_id
                    # Update cache
                    self._fk_caches.setdefault(cache_key, {})[lookup_val] = resolved_id
                else:
                    data[fm.target_field] = None

    def _resolve_implicit_fks(
        self,
        data: dict[str, Any],
        field_mappings: list[EtlFieldMapping],
    ):
        """Resolve implicit FK fields (boolean gate + code->UUID mappings).

        Handles two patterns:
        1. Boolean gate: x_to_bool=True + default_value -> FK lookup UUID
        2. Code mapping: map_sprache="de" -> UUID from bas_sprache.code
        """
        for fm in field_mappings:
            target = fm.target_field
            if target not in IMPLICIT_FK_MAP:
                continue
            value = data.get(target)
            if value is None:
                continue

            fk_spec = IMPLICIT_FK_MAP[target]
            table, field = fk_spec[0], fk_spec[1]
            fk_filter = fk_spec[2] if len(fk_spec) > 2 else None
            cache_key = f"{table}.{field}" + (f":{fk_filter}" if fk_filter else "")
            cache = self._fk_caches.get(cache_key, {})

            # Boolean gate: True -> use default_value as lookup key, False -> None
            if isinstance(value, bool):
                if value and fm.default_value:
                    lookup_val = fm.default_value
                else:
                    data[target] = None
                    continue
            else:
                lookup_val = str(value)

            resolved = cache.get(lookup_val)
            data[target] = str(resolved) if resolved else None

    # ============ Junction Table Processing ============

    async def _process_junction_row(
        self,
        row: dict[str, Any],
        jt_mapping: EtlTableMapping,
        jt_fields: list[EtlFieldMapping],
        ref_context: dict[str, Any],
        stats: dict,
        dry_run: bool = False,
    ):
        """Process a junction table mapping for one row (e.g., com_unternehmen_organisation)."""
        target_table = jt_mapping.target_table
        table_stats = stats.setdefault("junction_details", {}).setdefault(
            target_table, {"created": 0, "skipped": 0}
        )

        # Apply field mappings with ref_current context
        jt_data = self._apply_field_mappings(row, jt_fields, ref_context=ref_context)

        # Resolve FK transforms (both patterns)
        await self._resolve_fk_or_create(jt_data, jt_fields, row)
        await self._resolve_fk_lookups(jt_data, jt_fields)
        self._resolve_implicit_fks(jt_data, jt_fields)

        # Skip if any value is None (e.g., x_to_bool=False, no FK resolved)
        if any(v is None for v in jt_data.values()):
            stats["junction_skipped"] += 1
            table_stats["skipped"] += 1
            return

        if dry_run:
            stats["junction_created"] += 1
            table_stats["created"] += 1
            return

        # Check for existing record (dedup by FK reference fields only)
        dedup_cols = {k: v for k, v in jt_data.items() if k.endswith("_id")}
        if not dedup_cols:
            dedup_cols = jt_data  # Fallback: use all fields

        where_parts = []
        params = {}
        for col, val in dedup_cols.items():
            where_parts.append(f"{col} = :{col}")
            params[col] = val

        check_query = text(
            f"SELECT id FROM {target_table} WHERE {' AND '.join(where_parts)} LIMIT 1"
        )
        existing = await self.db.execute(check_query, params)
        if existing.fetchone():
            stats["junction_skipped"] += 1
            table_stats["skipped"] += 1
            return

        # Insert new junction record
        new_id = str(generate_uuid())
        now = datetime.utcnow()
        columns = ["id"] + list(jt_data.keys()) + ["erstellt_am"]
        placeholders = [":id"] + [f":{k}" for k in jt_data.keys()] + [":erstellt_am"]
        insert_params = dict(jt_data)
        insert_params["id"] = new_id
        insert_params["erstellt_am"] = now

        insert_query = text(
            f"INSERT INTO {target_table} ({', '.join(columns)}) "
            f"VALUES ({', '.join(placeholders)})"
        )
        await self.db.execute(insert_query, insert_params)
        stats["junction_created"] += 1
        table_stats["created"] += 1

        # Track junction import record
        self.db.add(EtlImportRecord(
            batch_id=self._batch_id,
            entity_type=f"junction:{target_table}",
            entity_id=new_id,
            action="created",
        ))

    # ============ Dedup Caches ============

    async def _build_dedup_caches(self):
        """Pre-load all existing Unternehmen into lookup caches."""
        # Load all non-deleted Unternehmen
        result = await self.db.execute(
            select(
                ComUnternehmen.id,
                ComUnternehmen.email,
                ComUnternehmen.website,
                ComUnternehmen.telefon,
                ComUnternehmen.firmierung,
                ComUnternehmen.strasse,
            ).where(ComUnternehmen.geloescht_am.is_(None))
        )
        for row in result.all():
            uid = str(row.id)
            if row.email:
                key = row.email.strip().lower()
                if key:
                    self._email_cache[key] = uid
            if row.website:
                key = _normalize_url_for_cache(row.website)
                if key:
                    self._url_cache[key] = uid
            if row.telefon:
                key = _normalize_phone_for_cache(row.telefon)
                if key:
                    self._phone_cache[key] = uid
            if row.firmierung:
                key = row.firmierung.strip().lower()
                if key:
                    self._name_cache[key] = uid

        # Load all external IDs
        result = await self.db.execute(
            select(ComExternalId).where(ComExternalId.entity_type == "unternehmen")
        )
        for eid in result.scalars().all():
            cache_key = f"{eid.source_name}:{eid.id_type}:{eid.external_value}"
            self._extid_cache[cache_key] = str(eid.entity_id)
            entity_key = f"{eid.entity_type}:{eid.entity_id}:{eid.source_name}:{eid.id_type}"
            self._entity_extid_set.add(entity_key)

        # Load business identifiers (USt-ID, DUNS, etc.)
        result = await self.db.execute(
            text("SELECT unternehmen_id, typ, wert FROM com_unternehmen_identifikation")
        )
        for row in result.all():
            key = f"{row[1]}:{row[2]}"
            self._ident_cache[key] = str(row[0])

        logger.info(
            f"Dedup caches loaded: {len(self._email_cache)} emails, "
            f"{len(self._url_cache)} urls, {len(self._phone_cache)} phones, "
            f"{len(self._name_cache)} names, {len(self._extid_cache)} external IDs, "
            f"{len(self._ident_cache)} business IDs"
        )

    async def _build_fk_caches(self, field_mappings: list[EtlFieldMapping]):
        """Pre-build FK lookup caches for fk_lookup, fk_lookup_or_create, and implicit FKs."""
        built = set()
        for fm in field_mappings:
            # Explicit fk_lookup: / fk_lookup_or_create: transforms
            if fm.transform:
                for prefix in ("fk_lookup:", "fk_lookup_or_create:"):
                    if fm.transform.startswith(prefix):
                        spec = fm.transform[len(prefix):]
                        if "." in spec:
                            table, field = spec.split(".", 1)
                            cache_key = f"{table}.{field}"
                            if cache_key not in built:
                                await self._etl_service.build_fk_lookup_cache(table, field)
                                if cache_key in self._etl_service._fk_cache:
                                    self._fk_caches[cache_key] = self._etl_service._fk_cache[cache_key]
                                built.add(cache_key)
                        break

            # Implicit FK fields (boolean gate, sprache_id, status_id, etc.)
            if fm.target_field in IMPLICIT_FK_MAP:
                fk_spec = IMPLICIT_FK_MAP[fm.target_field]
                table, field = fk_spec[0], fk_spec[1]
                fk_filter = fk_spec[2] if len(fk_spec) > 2 else None
                cache_key = f"{table}.{field}" + (f":{fk_filter}" if fk_filter else "")
                if cache_key not in built:
                    if fk_filter:
                        # Filtered lookup: build cache with WHERE clause
                        query = text(
                            f"SELECT {field}, id FROM {table} "
                            f"WHERE {field} IS NOT NULL AND {fk_filter}"
                        )
                        result = await self.db.execute(query)
                        self._fk_caches[cache_key] = {
                            row[0]: row[1] for row in result.fetchall()
                        }
                    else:
                        await self._etl_service.build_fk_lookup_cache(table, field)
                        if cache_key in self._etl_service._fk_cache:
                            self._fk_caches[cache_key] = self._etl_service._fk_cache[cache_key]
                    built.add(cache_key)

    async def _build_plz_cache(self):
        """Pre-load PLZ -> GeoOrt.id mapping."""
        result = await self.db.execute(
            text("""
                SELECT plz, id, ist_hauptort
                FROM geo_ort
                WHERE plz IS NOT NULL AND plz != ''
                ORDER BY ist_hauptort DESC
            """)
        )
        for row in result.all():
            plz = str(row[0]).strip().zfill(5)
            ort_id = str(row[1])
            # First match wins (ist_hauptort=True first due to ORDER BY)
            if plz not in self._plz_cache:
                self._plz_cache[plz] = ort_id

        logger.info(f"PLZ cache loaded: {len(self._plz_cache)} entries")

    def _update_dedup_caches(self, unternehmen_id: str, u_data: dict):
        """Update caches after creating/updating an Unternehmen."""
        uid = unternehmen_id
        if u_data.get("email"):
            self._email_cache[u_data["email"].strip().lower()] = uid
        if u_data.get("website"):
            key = _normalize_url_for_cache(u_data["website"])
            if key:
                self._url_cache[key] = uid
        if u_data.get("telefon"):
            key = _normalize_phone_for_cache(u_data["telefon"])
            if key:
                self._phone_cache[key] = uid
        if u_data.get("firmierung"):
            self._name_cache[u_data["firmierung"].strip().lower()] = uid

    # ============ Dedup Logic ============

    def _dedup_unternehmen(
        self,
        u_data: dict,
        source_row: dict,
        source_name: str,
        extid_fields: list[tuple[str, str, str]],
        bizid_fields: list[tuple[str, str]] | None = None,
    ) -> tuple[str | None, str | None]:
        """
        Multi-stage deduplication cascade.
        Returns: (existing_unternehmen_id, match_method)

        Priority order:
        1. External ID (Asana Task ID, Buschdata Kd-Nr.)
        2. Business Identifier (USt-ID, DUNS, etc.)
        3. Email
        4. Website
        5. Phone
        6. Firmierung
        """
        # Stage 1: External ID match
        for src_field, ext_source, ext_type, *rest in extid_fields:
            aliases = rest[0] if rest else None
            ext_val = self._get_value_with_aliases(source_row, src_field, aliases)
            if ext_val is not None and str(ext_val).strip():
                cache_key = f"{ext_source}:{ext_type}:{str(ext_val).strip()}"
                if cache_key in self._extid_cache:
                    return self._extid_cache[cache_key], "external_id"

        # Stage 2: Business Identifier match (USt-ID, DUNS, etc.)
        if bizid_fields:
            for src_field, ident_type, *rest in bizid_fields:
                aliases = rest[0] if rest else None
                val = self._get_value_with_aliases(source_row, src_field, aliases)
                if val is not None and str(val).strip():
                    cache_key = f"{ident_type}:{str(val).strip()}"
                    if cache_key in self._ident_cache:
                        return self._ident_cache[cache_key], f"business_id:{ident_type}"

        # Stage 3: Email match (skip generic domains like gmail.com)
        if u_data.get("email"):
            key = u_data["email"].strip().lower()
            domain = key.split("@")[-1] if "@" in key else ""
            if domain not in GENERIC_EMAIL_DOMAINS and key in self._email_cache:
                return self._email_cache[key], "email"

        # Stage 4: Website match
        if u_data.get("website"):
            key = _normalize_url_for_cache(u_data["website"])
            if key and key in self._url_cache:
                return self._url_cache[key], "website"

        # Stage 5: Phone match
        if u_data.get("telefon"):
            key = _normalize_phone_for_cache(u_data["telefon"])
            if key and key in self._phone_cache:
                return self._phone_cache[key], "telefon"

        # Stage 6: Firmierung match
        if u_data.get("firmierung"):
            key = u_data["firmierung"].strip().lower()
            if key in self._name_cache:
                return self._name_cache[key], "firmierung"

        return None, None

    # ============ Upsert Logic ============

    async def _upsert_unternehmen(
        self,
        existing_id: str | None,
        u_data: dict[str, Any],
        field_mappings: list[EtlFieldMapping],
    ) -> tuple[str, str]:
        """Create or update Unternehmen respecting update_rules.

        Returns: (unternehmen_id, "created" | "updated")
        """
        # Build update_rule map: target_field -> rule
        rules = {}
        for fm in field_mappings:
            rules[fm.target_field] = fm.update_rule or "always"

        if existing_id:
            # UPDATE existing
            result = await self.db.execute(
                select(ComUnternehmen).where(ComUnternehmen.id == existing_id)
            )
            unternehmen = result.scalar_one_or_none()
            if not unternehmen:
                # ID from cache but not in DB — create new
                return await self._create_unternehmen(u_data), "created"

            updated = False
            for field, value in u_data.items():
                if field in ("id", "erstellt_am", "aktualisiert_am", "geloescht_am"):
                    continue
                rule = rules.get(field, "always")
                current_value = getattr(unternehmen, field, None)

                if rule == "never":
                    continue
                elif rule == "if_empty":
                    if current_value is not None and str(current_value).strip():
                        continue

                if value != current_value:
                    setattr(unternehmen, field, value)
                    updated = True

            if updated:
                unternehmen.aktualisiert_am = datetime.utcnow()
                await self.db.flush()

            return str(unternehmen.id), "updated"
        else:
            # CREATE new
            new_id = await self._create_unternehmen(u_data)
            return new_id, "created"

    async def _create_unternehmen(self, u_data: dict[str, Any]) -> str:
        """Create a new Unternehmen record."""
        new_id = generate_uuid()
        unternehmen = ComUnternehmen(id=new_id, **u_data)
        self.db.add(unternehmen)
        await self.db.flush()
        return str(new_id)

    async def _upsert_kontakt(
        self,
        unternehmen_id: str,
        k_data: dict[str, Any],
        field_mappings: list[EtlFieldMapping],
    ) -> tuple[str, str]:
        """Create or update Kontakt. Dedup by email within Unternehmen.

        Returns: (kontakt_id, "created" | "updated")
        """
        # Try to find existing kontakt by email within this Unternehmen
        existing_kontakt = None
        if k_data.get("email"):
            result = await self.db.execute(
                select(ComKontakt)
                .where(ComKontakt.unternehmen_id == unternehmen_id)
                .where(ComKontakt.email == k_data["email"])
                .where(ComKontakt.geloescht_am.is_(None))
            )
            existing_kontakt = result.scalars().first()

        if existing_kontakt:
            # Update existing kontakt
            rules = {}
            for fm in field_mappings:
                rules[fm.target_field] = fm.update_rule or "always"

            updated = False
            for field, value in k_data.items():
                if field in ("id", "unternehmen_id", "erstellt_am", "aktualisiert_am"):
                    continue
                rule = rules.get(field, "always")
                current = getattr(existing_kontakt, field, None)

                if rule == "never":
                    continue
                elif rule == "if_empty":
                    if current is not None and str(current).strip():
                        continue

                if value != current:
                    setattr(existing_kontakt, field, value)
                    updated = True

            if updated:
                existing_kontakt.aktualisiert_am = datetime.utcnow()
                await self.db.flush()

            return str(existing_kontakt.id), "updated"
        else:
            # Create new kontakt
            new_id = generate_uuid()
            # Check if this is the first kontakt for the Unternehmen
            count_result = await self.db.execute(
                select(ComKontakt)
                .where(ComKontakt.unternehmen_id == unternehmen_id)
                .where(ComKontakt.geloescht_am.is_(None))
            )
            is_first = count_result.scalar_one_or_none() is None

            kontakt = ComKontakt(
                id=new_id,
                unternehmen_id=unternehmen_id,
                ist_hauptkontakt=is_first,
                **k_data,
            )
            self.db.add(kontakt)
            await self.db.flush()
            return str(new_id), "created"

    # ============ External IDs ============

    async def _save_external_ids(
        self,
        entity_type: str,
        entity_id: str,
        source_row: dict[str, Any],
        extid_fields: list[tuple],
    ):
        """Save external IDs from source row to com_external_id."""
        for src_field, ext_source, ext_type, *rest in extid_fields:
            aliases = rest[0] if rest else None
            ext_val = self._get_value_with_aliases(source_row, src_field, aliases)
            if ext_val is None or not str(ext_val).strip():
                continue

            ext_val_str = str(ext_val).strip()
            cache_key = f"{ext_source}:{ext_type}:{ext_val_str}"

            # Check if this exact value already exists
            if cache_key in self._extid_cache:
                continue

            # Check if entity already has an external_id for this source+type
            entity_key = f"{entity_type}:{entity_id}:{ext_source}:{ext_type}"
            if entity_key in self._entity_extid_set:
                continue

            # Create new external ID
            new_extid = ComExternalId(
                entity_type=entity_type,
                entity_id=entity_id,
                source_name=ext_source,
                id_type=ext_type,
                external_value=ext_val_str,
            )
            self.db.add(new_extid)
            self._extid_cache[cache_key] = entity_id
            self._entity_extid_set.add(entity_key)

    # ============ Business Identifiers ============

    async def _save_business_ids(
        self,
        unternehmen_id: str,
        source_row: dict[str, Any],
        bizid_fields: list[tuple],
    ):
        """Save business identifiers (USt-ID, DUNS, etc.) to com_unternehmen_identifikation."""
        for src_field, ident_type, *rest in bizid_fields:
            aliases = rest[0] if rest else None
            value = self._get_value_with_aliases(source_row, src_field, aliases)
            if value is None or not str(value).strip():
                continue

            value_str = str(value).strip()
            cache_key = f"{ident_type}:{value_str}"

            # Already exists for this or another Unternehmen?
            if cache_key in self._ident_cache:
                continue

            # Upsert: check if this Unternehmen already has this type
            existing = await self.db.execute(
                select(ComUnternehmenIdentifikation)
                .where(ComUnternehmenIdentifikation.unternehmen_id == unternehmen_id)
                .where(ComUnternehmenIdentifikation.typ == ident_type)
            )
            if existing.scalar_one_or_none():
                continue

            new_ident = ComUnternehmenIdentifikation(
                unternehmen_id=unternehmen_id,
                typ=ident_type,
                wert=value_str,
            )
            self.db.add(new_ident)
            self._ident_cache[cache_key] = unternehmen_id


# ============ Helper Functions ============

def _normalize_url_for_cache(url: str) -> str | None:
    """Normalize URL for cache lookup."""
    import re
    if not url or not url.strip():
        return None
    url = url.strip().lower()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = url.rstrip('/')
    return url if url else None


def _normalize_phone_for_cache(phone: str) -> str | None:
    """Normalize phone for cache lookup."""
    import re
    if not phone or not phone.strip():
        return None
    p = phone.strip()
    p = re.sub(r'^\+49\s*', '0', p)
    p = re.sub(r'^0049\s*', '0', p)
    p = re.sub(r'[^\d]', '', p)
    return p if p else None
