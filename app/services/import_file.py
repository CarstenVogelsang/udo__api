"""
Import File Service for centralized Excel file management.

Handles:
- File upload + persist to disk
- Excel header parsing + row count extraction
- Auto-matching against configured EtlSource field mappings
- Paginated row retrieval from stored files
- File assignment to EtlSource projects
"""
import json
import logging
import os
import uuid
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models.etl import EtlSource, EtlTableMapping, EtlFieldMapping, EtlImportFile

logger = logging.getLogger(__name__)


class ImportFileService:
    """Manages uploaded import files: storage, analysis, and retrieval."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self._settings = get_settings()

    # ============ Upload ============

    async def upload_file(
        self,
        content: bytes,
        filename: str,
        content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        uploaded_by: str | None = None,
    ) -> EtlImportFile:
        """Upload and analyze an Excel file.

        Saves file to disk, parses headers + row count, runs auto-matching
        against all active Excel sources.
        """
        stored_filename = f"{uuid.uuid4().hex}.xlsx"

        # Parse Excel metadata
        try:
            headers, row_count = self._parse_excel_metadata(content)
        except Exception as e:
            logger.exception(f"Failed to parse Excel file: {e}")
            import_file = EtlImportFile(
                original_filename=filename,
                stored_filename=stored_filename,
                file_size=len(content),
                content_type=content_type,
                status="error",
                uploaded_by=uploaded_by,
                notizen=f"Parse-Fehler: {str(e)}",
            )
            self.db.add(import_file)
            await self.db.flush()
            return import_file

        # Auto-match against configured sources
        analysis, best_source_id = await self._analyze_source_match(headers)

        # Determine status + assignment
        if best_source_id:
            status = "assigned"
            source_id = best_source_id
        elif analysis:
            status = "analyzed"
            source_id = None
        else:
            status = "pending"
            source_id = None

        # Save to disk
        storage_dir = self._get_storage_dir(source_id)
        os.makedirs(storage_dir, exist_ok=True)
        file_path = os.path.join(storage_dir, stored_filename)
        with open(file_path, "wb") as f:
            f.write(content)

        # Create DB record
        import_file = EtlImportFile(
            source_id=source_id,
            original_filename=filename,
            stored_filename=stored_filename,
            file_size=len(content),
            content_type=content_type,
            status=status,
            headers=json.dumps(headers),
            row_count=row_count,
            analysis_result=json.dumps(analysis) if analysis else None,
            uploaded_by=uploaded_by,
        )
        self.db.add(import_file)
        await self.db.flush()

        logger.info(
            f"Uploaded {filename}: {row_count} rows, {len(headers)} columns, "
            f"status={status}, source_id={source_id}"
        )
        return import_file

    # ============ Row Retrieval ============

    def get_file_rows(
        self,
        import_file: EtlImportFile,
        offset: int = 0,
        limit: int = 5000,
    ) -> tuple[list[str], list[dict[str, Any]]]:
        """Read rows from stored Excel file (paginated).

        Returns (headers, rows) where rows is a list of {column: value} dicts.
        """
        file_path = self._get_file_path(import_file)
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found on disk: {file_path}")

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows()
        header_row = next(rows_iter)
        headers = [
            cell.value.strip() if isinstance(cell.value, str) else str(cell.value or "")
            for cell in header_row
        ]

        result = []
        current_idx = 0
        for row in rows_iter:
            row_dict = {}
            for header, cell in zip(headers, row):
                if header:
                    value = cell.value
                    # Serialize datetime for JSON safety
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    row_dict[header] = value

            # Skip completely empty rows
            if not any(v is not None and str(v).strip() for v in row_dict.values()):
                continue

            if current_idx >= offset:
                result.append(row_dict)
                if len(result) >= limit:
                    break
            current_idx += 1

        wb.close()
        return headers, result

    # ============ Assignment ============

    async def assign_to_source(
        self, import_file: EtlImportFile, source_id: str
    ) -> EtlImportFile:
        """Assign (or reassign) an import file to an EtlSource.

        Moves the file on disk to the source-specific directory.
        """
        old_path = self._get_file_path(import_file)

        import_file.source_id = source_id
        import_file.status = "assigned"

        new_dir = self._get_storage_dir(source_id)
        os.makedirs(new_dir, exist_ok=True)
        new_path = os.path.join(new_dir, import_file.stored_filename)

        # Move file on disk (if it exists at old location)
        if os.path.exists(old_path) and old_path != new_path:
            os.rename(old_path, new_path)

        await self.db.flush()
        return import_file

    # ============ Deletion ============

    async def delete_file(self, import_file: EtlImportFile) -> None:
        """Delete import file from DB and disk."""
        file_path = self._get_file_path(import_file)

        # Delete from disk
        if os.path.exists(file_path):
            os.remove(file_path)

        # Delete from DB
        await self.db.delete(import_file)
        await self.db.flush()

    # ============ Internal: Excel Parsing ============

    def _parse_excel_metadata(self, content: bytes) -> tuple[list[str], int]:
        """Extract headers and row count from Excel content.

        Returns (headers, row_count).
        Uses ws.max_row for fast count instead of iterating all rows.
        """
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows()
        header_row = next(rows_iter)
        headers = [
            cell.value.strip() if isinstance(cell.value, str) else str(cell.value or "")
            for cell in header_row
        ]
        # Filter out empty header columns
        headers = [h for h in headers if h]

        # Fast row count: max_row - 1 (header row)
        # max_row is metadata from the xlsx file, no iteration needed
        row_count = max((ws.max_row or 1) - 1, 0)

        wb.close()
        return headers, row_count

    # ============ Internal: Source Matching ============

    async def _analyze_source_match(
        self, file_headers: list[str]
    ) -> tuple[list[dict[str, Any]], str | None]:
        """Match file headers against all active Excel sources (alias-aware).

        Returns (analysis_results, best_source_id_or_none).
        Considers both source_field and source_field_aliases for matching.
        Coverage denominator = canonical source_fields only (aliases don't inflate count).
        """
        result = await self.db.execute(
            select(EtlSource)
            .where(EtlSource.connection_type == "excel")
            .where(EtlSource.is_active.is_(True))
        )
        sources = result.scalars().all()

        if not sources:
            return [], None

        file_headers_lower = {h.lower().strip() for h in file_headers}
        analysis = []
        best_score = 0.0
        best_source_id = None

        for source in sources:
            source_fields = set()          # Canonical fields (coverage denominator)
            alias_to_canonical = {}        # alias.lower() -> canonical.lower()

            for tm in source.table_mappings:
                for fm in tm.field_mappings:
                    field = fm.source_field.strip()
                    if field.startswith("__"):
                        continue
                    canonical = field.lower()
                    source_fields.add(canonical)

                    # Parse aliases
                    if fm.source_field_aliases:
                        aliases = _parse_aliases(fm.source_field_aliases)
                        for alias in aliases:
                            alias_to_canonical[alias.strip().lower()] = canonical

            if not source_fields:
                continue

            # Match: file header hits source_field OR an alias
            matched_canonicals = set()
            for header in file_headers_lower:
                if header in source_fields:
                    matched_canonicals.add(header)
                elif header in alias_to_canonical:
                    matched_canonicals.add(alias_to_canonical[header])

            coverage = len(matched_canonicals) / len(source_fields)

            analysis.append({
                "source_id": str(source.id),
                "source_name": source.name,
                "source_titel": source.titel or source.description or source.name,
                "coverage": round(coverage, 3),
                "matched_fields": len(matched_canonicals),
                "total_fields": len(source_fields),
            })

            if coverage >= 0.8 and coverage > best_score:
                best_score = coverage
                best_source_id = str(source.id)

        # Sort by coverage descending
        analysis.sort(key=lambda x: x["coverage"], reverse=True)

        return analysis, best_source_id

    # ============ Compatibility Report ============

    async def compute_compatibility(
        self,
        import_file: EtlImportFile,
        source_id: str,
    ) -> dict[str, Any]:
        """Compute field-level compatibility between a file and a source.

        Returns a detailed report: matched/missing/unmapped fields with
        fuzzy suggestions for resolving mismatches.
        """
        import difflib

        # Parse file headers
        headers_raw = json.loads(import_file.headers) if isinstance(
            import_file.headers, str
        ) else (import_file.headers or [])
        file_headers = {h.strip() for h in headers_raw if h.strip()}
        # lower -> original mapping for display
        file_headers_lower = {h.lower(): h for h in file_headers}

        # Load source
        result = await self.db.execute(
            select(EtlSource).where(EtlSource.id == source_id)
        )
        source = result.scalar_one_or_none()
        if not source:
            raise ValueError(f"Source {source_id} not found")

        fields_report = []
        matched_headers = set()  # Track consumed file headers

        for tm in source.table_mappings:
            for fm in tm.field_mappings:
                if fm.source_field.strip().startswith("__"):
                    continue

                canonical = fm.source_field.strip()
                canonical_lower = canonical.lower()
                aliases = _parse_aliases(fm.source_field_aliases)

                # Try matching: 1) exact, 2) alias
                status = "missing"
                matched_header = None

                if canonical_lower in file_headers_lower:
                    status = "matched"
                    matched_header = file_headers_lower[canonical_lower]
                    matched_headers.add(matched_header)
                else:
                    for alias in aliases:
                        alias_lower = alias.strip().lower()
                        if alias_lower in file_headers_lower:
                            status = "matched_by_alias"
                            matched_header = file_headers_lower[alias_lower]
                            matched_headers.add(matched_header)
                            break

                # Fuzzy suggestions for missing fields
                suggestions = []
                if status == "missing":
                    available = [h for h in file_headers if h not in matched_headers]
                    for h in available:
                        sim = difflib.SequenceMatcher(
                            None, canonical_lower, h.lower()
                        ).ratio()
                        norm_sim = difflib.SequenceMatcher(
                            None,
                            _normalize_for_fuzzy(canonical_lower),
                            _normalize_for_fuzzy(h.lower()),
                        ).ratio()
                        best_sim = max(sim, norm_sim)
                        if best_sim >= 0.5:
                            suggestions.append({
                                "header": h,
                                "similarity": round(sim, 3),
                                "normalized_similarity": round(norm_sim, 3),
                            })
                    suggestions.sort(
                        key=lambda s: max(
                            s["similarity"], s["normalized_similarity"]
                        ),
                        reverse=True,
                    )
                    suggestions = suggestions[:3]

                fields_report.append({
                    "field_mapping_id": str(fm.id),
                    "source_field": canonical,
                    "target_field": fm.target_field,
                    "status": status,
                    "matched_header": matched_header,
                    "aliases": aliases,
                    "is_required": fm.is_required,
                    "suggestions": suggestions,
                })

        # Unmapped headers
        unmapped = []
        for h in file_headers:
            if h not in matched_headers:
                reverse_suggestions = []
                for entry in fields_report:
                    if entry["status"] == "missing":
                        sim = difflib.SequenceMatcher(
                            None, h.lower(), entry["source_field"].lower()
                        ).ratio()
                        if sim >= 0.4:
                            reverse_suggestions.append({
                                "field_mapping_id": entry["field_mapping_id"],
                                "source_field": entry["source_field"],
                                "similarity": round(sim, 3),
                            })
                reverse_suggestions.sort(
                    key=lambda s: s["similarity"], reverse=True
                )
                unmapped.append({
                    "header": h,
                    "suggestions": reverse_suggestions[:3],
                })

        matched_count = sum(1 for f in fields_report if f["status"] == "matched")
        alias_count = sum(
            1 for f in fields_report if f["status"] == "matched_by_alias"
        )
        missing_count = sum(1 for f in fields_report if f["status"] == "missing")

        return {
            "file_id": str(import_file.id),
            "source_id": str(source.id),
            "source_name": source.name,
            "total_fields": len(fields_report),
            "matched_count": matched_count,
            "matched_by_alias_count": alias_count,
            "missing_count": missing_count,
            "unmapped_count": len(unmapped),
            "coverage": round(
                (matched_count + alias_count) / max(len(fields_report), 1), 3
            ),
            "fields": fields_report,
            "unmapped_headers": unmapped,
        }

    # ============ Internal: File Paths ============

    def _get_storage_dir(self, source_id: str | None) -> str:
        """Get storage directory for a source (or 'unassigned')."""
        base = os.path.join(self._settings.upload_dir, "etl")
        if source_id:
            return os.path.join(base, str(source_id))
        return os.path.join(base, "unassigned")

    def _get_file_path(self, import_file: EtlImportFile) -> str:
        """Get full file path for an import file."""
        storage_dir = self._get_storage_dir(
            str(import_file.source_id) if import_file.source_id else None
        )
        return os.path.join(storage_dir, import_file.stored_filename)


# ============ Module-level Helpers ============

def _parse_aliases(aliases_value: str | list | None) -> list[str]:
    """Parse source_field_aliases from DB (JSON string or list)."""
    if not aliases_value:
        return []
    if isinstance(aliases_value, list):
        return aliases_value
    try:
        parsed = json.loads(aliases_value)
        return parsed if isinstance(parsed, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _normalize_for_fuzzy(text: str) -> str:
    """Normalize text for fuzzy comparison: umlauts, special chars."""
    result = text.lower()
    for old, new in (
        ("ä", "ae"), ("ö", "oe"), ("ü", "ue"), ("ß", "ss"),
        ("-", ""), ("_", ""), (" ", ""), (".", ""), ("/", ""),
        ("(", ""), (")", ""), ("\"", ""),
    ):
        result = result.replace(old, new)
    return result
